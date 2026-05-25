"""
import_to_mysql.py

Reads crm_sample.xlsx (or crm_template.xlsx for real data) from the
'anonymized_data' folder one level above this script and loads it into
the networking_crm MySQL database.

Handles the temp ID → real MySQL ID mapping automatically so foreign
key relationships are preserved correctly.

Insert order:
    Companies → Contacts → Interactions → Referrals

Folder structure expected:
    project/
    ├── anonymized_data/
    │   └── crm_sample.xlsx        ← anonymized data for testing
    ├── input/
    │   └── crm_template.xlsx      ← real data (personal use only)
    └── python/
        └── import_to_mysql.py     ← this script

Usage:
    uv run import_to_mysql.py
    # or
    python import_to_mysql.py

Requirements:
    uv add mysql-connector-python openpyxl python-dotenv
"""

import os
from pathlib import Path
import mysql.connector
from openpyxl import load_workbook
from datetime import datetime, date
from dotenv import load_dotenv

# ── Paths ────────────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent
INPUT_FILE = SCRIPT_DIR.parent / "anonymized_data" / "crm_anonymized_data.xlsx"

# ── Database config ───────────────────────────────────────────────────────────
# Credentials are loaded from a .env file at the project root.
# Never hardcode passwords here — the .env file must be in .gitignore.
#
# .env file should contain:
#   DB_HOST=localhost
#   DB_USER=root
#   DB_PASSWORD=your_password_here
#   DB_NAME=networking_crm

load_dotenv(SCRIPT_DIR.parent / ".env")

DB_CONFIG = {
    "host":     os.getenv("DB_HOST", "localhost"),
    "user":     os.getenv("DB_USER", "root"),
    "password": os.getenv("DB_PASSWORD", ""),
    "database": os.getenv("DB_NAME", "networking_crm"),
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def get_headers(ws) -> dict:
    """Return a dict of { column_name: column_index (1-based) } from row 1."""
    return {
        cell.value: cell.column
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
        if cell.value is not None
    }


def clean_date(value):
    """Normalize date values from Excel to a MySQL-compatible date string or None."""
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str) and value.strip():
        return value.strip()
    return None


def clean_str(value):
    """Strip strings and return None for empty values."""
    if value is None:
        return None
    stripped = str(value).strip()
    return stripped if stripped else None


def rows_as_dicts(ws) -> list[dict]:
    """Return all non-empty data rows as a list of dicts keyed by header name."""
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        results.append(dict(zip(headers, row)))
    return results


# ── Import functions ──────────────────────────────────────────────────────────

def import_companies(cursor, rows: list[dict]) -> dict:
    """
    Insert companies and return a mapping of temp Excel ID → real MySQL ID.

    Returns:
        dict: { temp_id (int): real_mysql_id (int) }
    """
    id_map = {}
    sql = """
        INSERT INTO companies (name, industry, website, notes)
        VALUES (%s, %s, %s, %s)
    """
    for row in rows:
        temp_id = row.get("id")
        if temp_id is None:
            continue

        values = (
            clean_str(row.get("name")),
            clean_str(row.get("industry")),
            clean_str(row.get("website")),
            clean_str(row.get("notes")),
        )
        cursor.execute(sql, values)
        id_map[int(temp_id)] = cursor.lastrowid

    print(f"  Inserted {len(id_map)} companies")
    return id_map


def import_contacts(cursor, rows: list[dict], company_id_map: dict) -> dict:
    """
    Insert contacts, translating temp company IDs to real MySQL IDs.

    Returns:
        dict: { temp_id (int): real_mysql_id (int) }
    """
    id_map = {}
    sql = """
        INSERT INTO contacts (
            company_id, first_name, last_name, email, phone,
            linkedin_url, location, how_we_met, met_on
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    for row in rows:
        temp_id = row.get("id")
        if temp_id is None:
            continue

        temp_company_id = row.get("company_id")
        real_company_id = company_id_map.get(int(temp_company_id)) if temp_company_id else None

        values = (
            real_company_id,
            clean_str(row.get("first_name")),
            clean_str(row.get("last_name")),
            clean_str(row.get("email")),
            clean_str(row.get("phone")),
            clean_str(row.get("linkedin_url")),
            clean_str(row.get("location")),
            clean_str(row.get("how_we_met")),
            clean_date(row.get("met_on")),
        )
        cursor.execute(sql, values)
        id_map[int(temp_id)] = cursor.lastrowid

    print(f"  Inserted {len(id_map)} contacts")
    return id_map


def import_interactions(cursor, rows: list[dict], contact_id_map: dict):
    """Insert interactions, translating temp contact IDs to real MySQL IDs."""
    sql = """
        INSERT INTO interactions (contact_id, date, type, summary, outcome)
        VALUES (%s, %s, %s, %s, %s)
    """
    count = 0
    for row in rows:
        temp_contact_id = row.get("contact_id")
        if temp_contact_id is None:
            continue

        real_contact_id = contact_id_map.get(int(temp_contact_id))
        if real_contact_id is None:
            print(f"  WARNING: contact_id {temp_contact_id} not found — skipping interaction row")
            continue

        values = (
            real_contact_id,
            clean_date(row.get("date")),
            clean_str(row.get("type")),
            clean_str(row.get("summary")),
            clean_str(row.get("outcome")),
        )
        cursor.execute(sql, values)
        count += 1

    print(f"  Inserted {count} interactions")


def import_referrals(cursor, rows: list[dict], contact_id_map: dict):
    """Insert referrals, translating both referrer and referred temp IDs."""
    sql = """
        INSERT INTO referrals (referrer_id, referred_id)
        VALUES (%s, %s)
    """
    count = 0
    for row in rows:
        temp_referrer_id = row.get("referrer_id")
        temp_referred_id = row.get("referred_id")

        if temp_referrer_id is None or temp_referred_id is None:
            continue

        real_referrer_id = contact_id_map.get(int(temp_referrer_id))
        real_referred_id = contact_id_map.get(int(temp_referred_id))

        if real_referrer_id is None or real_referred_id is None:
            print(f"  WARNING: referrer_id {temp_referrer_id} or referred_id {temp_referred_id} not found — skipping")
            continue

        values = (
            real_referrer_id,
            real_referred_id,
        )
        cursor.execute(sql, values)
        count += 1

    print(f"  Inserted {count} referrals")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"Input file not found: {INPUT_FILE}")

    print(f"Reading {INPUT_FILE.name}...")
    wb = load_workbook(INPUT_FILE)

    # Print actual headers found in each sheet for diagnosis
    for sheet_name in ["Companies", "Contacts", "Interactions", "Referrals"]:
        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        print(f"  {sheet_name} headers: {headers}")

    companies_rows    = rows_as_dicts(wb["Companies"])
    contacts_rows     = rows_as_dicts(wb["Contacts"])
    interactions_rows = rows_as_dicts(wb["Interactions"])
    referrals_rows    = rows_as_dicts(wb["Referrals"])

    print(f"  Found {len(companies_rows)} companies, {len(contacts_rows)} contacts, "
          f"{len(interactions_rows)} interactions, {len(referrals_rows)} referrals")

    print("\nConnecting to MySQL...")
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()
    print("  Connected")

    try:

        print("Clearing existing data...")
        cursor.execute("SET FOREIGN_KEY_CHECKS = 0;")
        cursor.execute("TRUNCATE TABLE referrals")
        cursor.execute("TRUNCATE TABLE interactions")
        cursor.execute("TRUNCATE TABLE contacts")
        cursor.execute("TRUNCATE TABLE companies")
        cursor.execute("SET FOREIGN_KEY_CHECKS = 1")

        print("\nInserting companies...")
        company_id_map = import_companies(cursor, companies_rows)

        print("Inserting contacts...")
        contact_id_map = import_contacts(cursor, contacts_rows, company_id_map)

        print("Inserting interactions...")
        import_interactions(cursor, interactions_rows, contact_id_map)

        print("Inserting referrals...")
        import_referrals(cursor, referrals_rows, contact_id_map)

        conn.commit()
        print("\nAll data committed successfully.")

    except Exception as e:
        conn.rollback()
        print(f"\nERROR: {e}")
        print("Transaction rolled back — no data was written to the database.")
        print("\nDIAGNOSTIC — first row of each sheet:")
        for sheet_name, rows in [
            ("Companies", companies_rows),
            ("Contacts", contacts_rows),
            ("Interactions", interactions_rows),
            ("Referrals", referrals_rows),
        ]:
            if rows:
                print(f"  {sheet_name}: {rows[0]}")
        raise

    finally:
        cursor.close()
        conn.close()
        print("Connection closed.")


if __name__ == "__main__":
    main()
