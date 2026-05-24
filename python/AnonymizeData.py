"""
anonymize.py

Reads crm_template.xlsx and produces crm_sample.xlsx with all
personally identifiable information replaced by Faker-generated values.

Sensitive fields replaced:
  Contacts  : first_name, last_name, email, phone, linkedin_url, location, notes
  Companies : name, website, notes  (industry/size/status kept)
  Interactions: summary, outcome
  Referrals : notes

Non-sensitive fields kept as-is:
  All IDs and foreign keys, dates, status fields, tags, industry,
  company size, interaction type, how_we_met, outreach_status

Usage:
    uv run anonymize.py

Output:
    crm_anonymized_data  (safe to commit to GitHub)
"""

from faker import Faker
from openpyxl import load_workbook
from pathlib import Path
from datetime import date

Faker.seed(1228673)

INPUT_PATH = Path("real_data")
INPUT_FILE = INPUT_PATH / "crm_real_data.xlsx"

OUTPUT_PATH = Path('anonymized_data/')
OUTPUT_PATH.mkdir(parents = True, exist_ok= True)
OUTPUT_FILE = OUTPUT_PATH / "crm_anonymized_data.xlsx"

fake = Faker()


def build_company_map(companies_sheet) -> dict:
    """
    Build a mapping of company temp ID -> fake company name and website.

    Returns:
        dict: { company_id (int) : { name, website } }
    """
    company_map = {}

    headers = [cell.value for cell in next(companies_sheet.iter_rows(min_row=1, max_row=1))]
    id_col = headers.index("id")

    for row in companies_sheet.iter_rows(min_row=2, values_only=True):
        company_id = row[id_col]
        if company_id is None:
            continue

        company_name = fake.company()
        domain = company_name.lower().replace(" ", "").replace(",", "").replace(".", "")[:15]
        company_map[company_id] = {
            "name": company_name,
            "website": f"https://www.{domain}.com",
        }

    return company_map


def build_contact_map(contacts_sheet) -> dict:
    """
    Build a mapping of contact temp ID -> fake identity.
    This ensures the same fake name is used consistently
    across Contacts, Interactions, and Referrals sheets.

    Returns:
        dict: { contact_id (int) : { first_name, last_name, email, phone, linkedin_url } }
    """
    contact_map = {}

    headers = [cell.value for cell in next(contacts_sheet.iter_rows(min_row=1, max_row=1))]
    id_col = headers.index("id")

    for row in contacts_sheet.iter_rows(min_row=2, values_only=True):
        contact_id = row[id_col]
        if contact_id is None:
            continue

        first = fake.first_name()
        last = fake.last_name()
        contact_map[contact_id] = {
            "first_name": first,
            "last_name": last,
            "email": fake.email(),
            "phone": fake.numerify('##########'),
            "linkedin_url": f"https://linkedin.com/in/{first.lower()}-{last.lower()}-{fake.bothify('####')}",
        }

    return contact_map

## anonymization

def anonymize_companies(ws, company_map: dict):
    """Replace name, website, industry, and notes in the Companies sheet."""
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    industries = ['Technology', 'Healthcare', 'Finance', 'Manufacturing', 'Retail']

    col = {name: idx + 1 for idx, name in enumerate(headers)}

    for row_idx in range(2, ws.max_row + 1):
        company_id = ws.cell(row=row_idx, column=col["id"]).value
        if company_id is None:
            continue

        fake_company = company_map.get(company_id, {})

        if "name" in col:
            ws.cell(row=row_idx, column=col["name"]).value = fake_company.get("name", fake.company())
        if "website" in col:
            ws.cell(row=row_idx, column=col["website"]).value = fake_company.get("website", "")
        if "industry" in col:
            ws.cell(row=row_idx, column=col["industry"]).value = fake.random_element(industries)
        if "notes" in col:
            ws.cell(row=row_idx, column=col["notes"]).value = fake.sentence() if ws.cell(row=row_idx, column=col["notes"]).value else None   


def anonymize_contacts(ws, contact_map: dict):
    """Replace first_name, last_name, email, phone, linkedin_url, location, notes in Contacts sheet."""
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    how_we_met = ['LinkedIn', 'college', 'life', 'referral', 'cold outreach', 'other']

    col = {name: idx + 1 for idx, name in enumerate(headers)}

    for row_idx in range(2, ws.max_row + 1):
        contact_id = ws.cell(row=row_idx, column=col["id"]).value
        if contact_id is None:
            continue

        fake_contact = contact_map.get(contact_id, {})

        if "first_name" in col:
            ws.cell(row=row_idx, column=col["first_name"]).value = fake_contact.get("first_name", fake.first_name())
        if "last_name" in col:
            ws.cell(row=row_idx, column=col["last_name"]).value = fake_contact.get("last_name", fake.last_name())
        if "email" in col:
            ws.cell(row=row_idx, column=col["email"]).value = fake_contact.get("email", fake.email())
        if "phone" in col:
            ws.cell(row=row_idx, column=col["phone"]).value = fake_contact.get("phone", fake.numerify('##########'))
        if "linkedin_url" in col:
            ws.cell(row=row_idx, column=col["linkedin_url"]).value = fake_contact.get("linkedin_url", "")
        if "location" in col:
            ws.cell(row=row_idx, column=col["location"]).value = fake_contact.get("location", fake.city())
        if "how_we_met" in col:
            ws.cell(row=row_idx, column=col["how_we_met"]).value = fake.random_element(how_we_met)
        if "notes" in col:
            ws.cell(row=row_idx, column=col["notes"]).value = fake.sentence() if ws.cell(row=row_idx, column=col["notes"]).value else None


def anonymize_interactions(ws):
    """Replace summary and outcome in Interactions sheet."""
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    interaction_type = ['email','call','in person', 'LinkedIn', 'thank you', 'other']
    col = {name: idx + 1 for idx, name in enumerate(headers)}

    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=col["id"]).value is None:
            continue

        if "summary" in col:
            ws.cell(row=row_idx, column=col["summary"]).value = fake.sentence() if ws.cell(row=row_idx, column=col["summary"]).value else None
        if "date" in col:
            ws.cell(row=row_idx, column=col["date"]).value = fake.date_between(start_date=date(2025, 1, 1), end_date=date(2026, 12, 31))
        if "type" in col:
            ws.cell(row=row_idx, column=col["type"]).value = fake.random_element(interaction_type)
        if "outcome" in col:
            ws.cell(row=row_idx, column=col["outcome"]).value = fake.sentence() if ws.cell(row=row_idx, column=col["outcome"]).value else None


def anonymize_referrals(ws):
    """Replace notes in Referrals sheet. IDs are kept for relational integrity."""
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx + 1 for idx, name in enumerate(headers)}

    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=col["id"]).value is None:
            continue

        if "notes" in col:
            ws.cell(row=row_idx, column=col["notes"]).value = fake.sentence() if ws.cell(row=row_idx, column=col["notes"]).value else None

## running everything

def main():
    print(f"Reading {INPUT_FILE}...")
    wb = load_workbook(INPUT_FILE)

    # Build identity maps from original data before any changes
    print("Building identity maps...")
    contact_map = build_contact_map(wb["Contacts"])
    company_map = build_company_map(wb["Companies"])

    print(f"  {len(contact_map)} contacts mapped")
    print(f"  {len(company_map)} companies mapped")

    # Anonymize each sheet
    print("Anonymizing Companies...")
    anonymize_companies(wb["Companies"], company_map)

    print("Anonymizing Contacts...")
    anonymize_contacts(wb["Contacts"], contact_map)

    print("Anonymizing Interactions...")
    anonymize_interactions(wb["Interactions"])

    print("Anonymizing Referrals...")
    anonymize_referrals(wb["Referrals"])

    # Save to new file; never overwrite the original
    wb.save(OUTPUT_FILE)
    print(f"\nDone. Anonymized file saved to: {OUTPUT_FILE}")
    print("This file is safe to commit to GitHub.")

if __name__ == "__main__":
    main()